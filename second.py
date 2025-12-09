import csv
from datetime import datetime
from openpyxl import Workbook

try:
    with open('people_data.csv', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        data = list(reader)
except FileNotFoundError:
    print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV")
    exit()

try:
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    sheet_names = ['all', 'younger_18', '18-45', '45-70', 'older_70']
    sheets = {name: wb.create_sheet(name) for name in sheet_names}

    counters = {name: 1 for name in sheet_names}

    header = ['№', 'Прізвище', "Ім'я", 'По батькові', 'Дата народження', 'Вік']
    for s in sheets.values():
        s.append(header)

    today = datetime.now()

    for p in data:
        try:
            bd = datetime.strptime(p['Дата народження'], '%d.%m.%Y')
            age = today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))
        except ValueError:
            continue 

        base_row = [0, p['Прізвище'], p["Ім'я"], p['По батькові'], p['Дата народження'], age]

        row_all = base_row[:]
        row_all[0] = counters['all']
        sheets['all'].append(row_all)
        counters['all'] += 1

        if age < 18:
            cat = 'younger_18'
        elif age <= 45:
            cat = '18-45'
        elif age <= 70:
            cat = '45-70'
        else:
            cat = 'older_70'

        row_cat = base_row[:]
        row_cat[0] = counters[cat]
        sheets[cat].append(row_cat)
        counters[cat] += 1

    wb.save('people_data.xlsx')
    print("Ok")

except Exception as e:
    print("Повідомлення про неможливість створення XLSX файлу")